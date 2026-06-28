#!/usr/bin/env python3
"""
Profetch — Prom XLS Generator (нова програма)
==============================================
Порівнює товари у вашому фіді з товарами на Пром через API,
генерує XLS файл у форматі Пром тільки з новими товарами
(включно з характеристиками), беручи дані з резервної копії Прому.

Використання:
  pip install requests openpyxl
  python prom_xls_new_products.py
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os
import sys
import time
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ──────────────────────────────────────────────
#  КОЛЬОРИ
# ──────────────────────────────────────────────
BG       = "#0f1117"
BG2      = "#1a1d27"
BG3      = "#22263a"
ACCENT   = "#4f8ef7"
SUCCESS  = "#3ecf8e"
WARNING  = "#f7b731"
DANGER   = "#fc5c65"
TEXT     = "#e8eaf6"
TEXT_DIM = "#6b7280"
BORDER   = "#2d3250"

FONT_TITLE = ("Segoe UI", 13, "bold")
FONT_LABEL = ("Segoe UI", 9)
FONT_SMALL = ("Segoe UI", 8)
FONT_MONO  = ("Consolas", 9)
FONT_BIG   = ("Segoe UI", 22, "bold")
FONT_BTN   = ("Segoe UI", 10, "bold")


# ──────────────────────────────────────────────
#  ЛОГІКА
# ──────────────────────────────────────────────

def get_prom_products(token, on_progress):
    """Завантажує всі артикули та зовнішні ID з Пром через API."""
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    prom_skus    = set()
    prom_ext_ids = set()
    limit   = 100
    last_id = None
    page    = 1

    while True:
        params = {"limit": limit, "status": "on_display,draft,deleted"}
        if last_id:
            params["last_id"] = last_id

        resp = requests.get(
            "https://my.prom.ua/api/v1/products/list",
            headers=headers, params=params, timeout=30
        )
        resp.raise_for_status()
        data     = resp.json()
        products = data.get("products", [])
        if not products:
            break

        for p in products:
            sku    = str(p.get("sku", "")).strip()
            ext_id = str(p.get("external_id", "")).strip()
            if sku:
                prom_skus.add(sku)
            if ext_id and ext_id not in ("None", ""):
                prom_ext_ids.add(ext_id)

        on_progress(f"   Сторінка {page}: {len(prom_skus)} артикулів завантажено")
        if len(products) < limit:
            break
        last_id = products[-1]["id"]
        page   += 1
        time.sleep(0.4)

    return prom_skus, prom_ext_ids


def read_prom_backup(xlsx_path, on_progress):
    """
    Читає резервну копію Прому (XLS експорт).
    Повертає:
      - headers: список заголовків (рядок 1)
      - rows: список кортежів значень (рядки 2+)
      - sku_index: індекс колонки Код_товара
      - ext_id_index: індекс колонки Ідентифікатор_товара
    """
    on_progress("   Читання резервної копії Пром...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    # Знаходимо ключові колонки
    sku_index    = next((i for i, h in enumerate(headers) if h == "Код_товара"), None)
    ext_id_index = next((i for i, h in enumerate(headers) if h == "Ідентифікатор_товара"
                         or h == "Идентификатор_товара"), None)

    if sku_index is None:
        raise ValueError("Колонка 'Код_товара' не знайдена у файлі!")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(row)

    on_progress(f"   Прочитано {len(rows)} товарів з резервної копії")
    wb.close()
    return headers, rows, sku_index, ext_id_index


def find_new_products(backup_rows, backup_sku_idx, backup_ext_idx, prom_skus, prom_ext_ids, on_progress):
    """
    Повертає тільки рядки з резервної копії, яких НЕ має на Пром через API.
    Це нові товари які треба додати.
    """
    on_progress("   Порівняння товарів...")
    new_rows = []
    already  = 0

    for row in backup_rows:
        sku    = str(row[backup_sku_idx]).strip() if backup_sku_idx is not None and row[backup_sku_idx] else ""
        ext_id = str(row[backup_ext_idx]).strip() if backup_ext_idx is not None and backup_ext_idx < len(row) and row[backup_ext_idx] else ""

        on_prom = (sku in prom_skus) or (ext_id in prom_ext_ids) or (sku in prom_ext_ids)

        if on_prom:
            already += 1
        else:
            new_rows.append(row)

    return new_rows, already


def generate_prom_xls(headers, new_rows, output_path, on_progress):
    """
    Генерує XLS файл у форматі Пром з новими товарами.
    Копіює структуру заголовків з резервної копії.
    """
    on_progress("   Генерація XLS файлу...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Export Products Sheet"

    # Стиль заголовка — копіюємо вигляд Прому
    header_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Рядок заголовків
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 32

    # Стиль рядків даних
    data_font  = Font(name="Arial", size=9)
    data_align = Alignment(vertical="top", wrap_text=False)

    # Додаємо нові товари
    for i, row in enumerate(new_rows):
        # Конвертуємо в список щоб можна редагувати
        row_data = list(row)
        # Заповнюємо None пустими рядками для кращої сумісності
        while len(row_data) < len(headers):
            row_data.append(None)
        ws.append(row_data)

        # Форматуємо рядок
        excel_row = i + 2
        for cell in ws[excel_row]:
            cell.font      = data_font
            cell.alignment = data_align

        ws.row_dimensions[excel_row].height = 20

    # Ширина колонок
    col_widths = {
        1: 12,   # Код_товара
        2: 50,   # Название
        3: 50,   # Название укр
        4: 30,   # Поисковые запросы
        5: 30,   # Поисковые запросы укр
        6: 60,   # Описание
        7: 60,   # Описание укр
        8: 10,   # Тип
        9: 10,   # Цена
        10: 8,   # Валюта
        15: 80,  # Ссылки изображений
        16: 10,  # Наличие
        26: 15,  # Идентификатор
        29: 15,  # Производитель
        30: 12,  # Страна
        43: 15,  # GTIN
        44: 15,  # MPN
    }
    for col_num, width in col_widths.items():
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = width

    # Заморожуємо перший рядок
    ws.freeze_panes = "A2"

    wb.save(output_path)
    size_kb = os.path.getsize(output_path) / 1024
    on_progress(f"   Збережено: {os.path.basename(output_path)} ({size_kb:.0f} КБ)")


# ──────────────────────────────────────────────
#  GUI
# ──────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Profetch — Prom XLS Generator")
        self.geometry("700x740")
        self.minsize(640, 660)
        self.configure(bg=BG)
        self.resizable(True, True)

        self.backup_path = tk.StringVar(value="")
        self.prom_token  = tk.StringVar(value="")
        self.out_path    = tk.StringVar(value="")
        self._running    = False

        self._build_ui()
        self._load_settings()

    def _build_ui(self):
        # Заголовок
        hdr = tk.Frame(self, bg=BG, pady=18)
        hdr.pack(fill="x", padx=24)
        tk.Label(hdr, text="⬡", font=("Segoe UI", 20), fg=ACCENT, bg=BG).pack(side="left")
        tk.Label(hdr, text="  PROFETCH", font=FONT_TITLE, fg=TEXT, bg=BG).pack(side="left")
        tk.Label(hdr, text="  XLS Generator", font=("Segoe UI", 10), fg=TEXT_DIM, bg=BG).pack(side="left", pady=3)

        self._sep()

        # БЛОК 1: Резервна копія
        self._section("1  Резервна копія з Пром.юа (.xlsx)")
        tk.Label(self,
                 text="Пром → Товари → Імпорт/Експорт → Експорт → завантажити файл",
                 font=FONT_SMALL, fg=TEXT_DIM, bg=BG).pack(anchor="w", padx=28, pady=(0, 4))

        r1 = tk.Frame(self, bg=BG)
        r1.pack(fill="x", padx=24, pady=(0, 8))
        e1 = self._entry(r1, self.backup_path, width=52)
        e1.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._btn_small(r1, "Огляд…", self._pick_backup).pack(side="left")

        # БЛОК 2: API токен
        self._section("2  API-токен Пром.юа")
        tk.Label(self,
                 text="Пром → Налаштування → Управління API-токенами",
                 font=FONT_SMALL, fg=TEXT_DIM, bg=BG).pack(anchor="w", padx=28, pady=(0, 4))

        r2 = tk.Frame(self, bg=BG)
        r2.pack(fill="x", padx=24, pady=(0, 8))
        self.tok_entry = self._entry(r2, self.prom_token, width=52, show="•")
        self.tok_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._btn_small(r2, "👁", self._toggle_token).pack(side="left")

        # БЛОК 3: Результат
        self._section("3  Зберегти результат як")
        r3 = tk.Frame(self, bg=BG)
        r3.pack(fill="x", padx=24, pady=(0, 8))
        e3 = self._entry(r3, self.out_path, width=52)
        e3.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._btn_small(r3, "Огляд…", self._pick_out).pack(side="left")

        self._sep()

        # Кнопка запуску
        self.run_btn = tk.Button(
            self, text="▶   ЗНАЙТИ НОВІ ТОВАРИ → XLS",
            font=FONT_BTN, bg=ACCENT, fg="#ffffff",
            activebackground="#7c5cfc", activeforeground="#ffffff",
            bd=0, cursor="hand2", padx=28, pady=12,
            command=self._start
        )
        self.run_btn.pack(fill="x", padx=24, pady=(12, 4))

        # Прогрес-бар
        pb_frame = tk.Frame(self, bg=BG, height=6)
        pb_frame.pack(fill="x", padx=24, pady=(0, 8))
        pb_frame.pack_propagate(False)
        self.pb_bg = tk.Frame(pb_frame, bg=BG3, height=6)
        self.pb_bg.pack(fill="both", expand=True)
        self.pb_fill = tk.Frame(self.pb_bg, bg=ACCENT, height=6, width=0)
        self.pb_fill.place(x=0, y=0, height=6)
        self._pb_anim = 0
        self._pb_dir  = 1

        # Лог
        tk.Label(self, text="Журнал", font=FONT_SMALL, fg=TEXT_DIM, bg=BG).pack(anchor="w", padx=24)
        self.log_box = tk.Text(
            self, font=FONT_MONO, bg=BG2, fg=TEXT,
            insertbackground=TEXT, bd=0, pady=10, padx=12,
            relief="flat", state="disabled", height=8,
            selectbackground=BG3
        )
        self.log_box.pack(fill="both", expand=True, padx=24, pady=(4, 0))

        # Статистика
        stat_frame = tk.Frame(self, bg=BG3, pady=14)
        stat_frame.pack(fill="x", padx=24, pady=8)

        self.stat_backup = self._stat_cell(stat_frame, "У резервній копії", "—")
        self.stat_prom   = self._stat_cell(stat_frame, "Вже на Пром",       "—")
        self.stat_new    = self._stat_cell(stat_frame, "НОВИХ → XLS",       "—", color=SUCCESS)

        for col in range(3):
            stat_frame.columnconfigure(col, weight=1)

        # Кнопка відкрити папку
        self.open_btn = tk.Button(
            self, text="📂  Відкрити папку з результатом",
            font=FONT_SMALL, bg=BG2, fg=TEXT_DIM,
            activebackground=BG3, activeforeground=TEXT,
            bd=0, cursor="hand2", padx=16, pady=8,
            command=self._open_folder, state="disabled"
        )
        self.open_btn.pack(pady=(0, 16))

    # ── Хелпери ──────────────────────────────

    def _sep(self):
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x", padx=24, pady=8)

    def _section(self, text):
        tk.Label(self, text=text, font=("Segoe UI", 9, "bold"),
                 fg=ACCENT, bg=BG).pack(anchor="w", padx=24, pady=(12, 4))

    def _entry(self, parent, var, width=40, show=None):
        kw = dict(textvariable=var, font=FONT_LABEL, bg=BG3, fg=TEXT,
                  insertbackground=TEXT, bd=0, relief="flat",
                  highlightthickness=1, highlightbackground=BORDER,
                  highlightcolor=ACCENT, width=width)
        if show:
            kw["show"] = show
        e = tk.Entry(parent, **kw)
        self._enable_paste(e)
        return e

    def _enable_paste(self, entry):
        """Вставка через праву кнопку миші та Ctrl+V незалежно від розкладки."""
        def do_paste(event=None):
            try:
                text = self.clipboard_get()
                try:
                    entry.delete("sel.first", "sel.last")
                except tk.TclError:
                    pass
                entry.insert("insert", text)
            except tk.TclError:
                pass
            return "break"

        def do_copy(event=None):
            try:
                self.clipboard_clear()
                self.clipboard_append(entry.selection_get())
            except tk.TclError:
                pass
            return "break"

        def do_cut(event=None):
            do_copy()
            try:
                entry.delete("sel.first", "sel.last")
            except tk.TclError:
                pass
            return "break"

        def select_all(event=None):
            entry.select_range(0, "end")
            entry.icursor("end")
            return "break"

        menu = tk.Menu(entry, tearoff=0, bg=BG3, fg=TEXT,
                       activebackground=ACCENT, activeforeground="#ffffff", bd=0)
        menu.add_command(label="Вставити", command=do_paste)
        menu.add_command(label="Копіювати", command=do_copy)
        menu.add_command(label="Вирізати", command=do_cut)
        menu.add_separator()
        menu.add_command(label="Виділити все", command=select_all)

        def show_menu(event):
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        entry.bind("<Button-3>", show_menu)
        entry.bind("<Control-v>", do_paste)
        entry.bind("<Control-V>", do_paste)
        entry.bind("<Control-c>", do_copy)
        entry.bind("<Control-C>", do_copy)
        entry.bind("<Control-x>", do_cut)
        entry.bind("<Control-X>", do_cut)
        entry.bind("<Control-a>", select_all)
        entry.bind("<Control-A>", select_all)

        def keycode_handler(event):
            if event.state & 0x4:
                if event.keycode == 86:
                    return do_paste()
                if event.keycode == 67:
                    return do_copy()
                if event.keycode == 88:
                    return do_cut()
                if event.keycode == 65:
                    return select_all()
        entry.bind("<Key>", keycode_handler)

    def _btn_small(self, parent, text, cmd):
        return tk.Button(parent, text=text, font=FONT_SMALL,
                         bg=BG3, fg=TEXT_DIM,
                         activebackground=BORDER, activeforeground=TEXT,
                         bd=0, cursor="hand2", padx=10, pady=6,
                         relief="flat", command=cmd,
                         highlightthickness=1, highlightbackground=BORDER)

    def _stat_cell(self, parent, label, value, color=TEXT):
        frame = tk.Frame(parent, bg=BG3)
        frame.grid(row=0, column=len(parent.winfo_children()) - 1,
                   sticky="nsew", padx=6)
        tk.Label(frame, text=label, font=FONT_SMALL, fg=TEXT_DIM, bg=BG3).pack(pady=(8, 0))
        val = tk.Label(frame, text=value, font=FONT_BIG, fg=color, bg=BG3)
        val.pack(pady=(0, 8))
        return val

    def _toggle_token(self):
        cur = self.tok_entry.cget("show")
        self.tok_entry.configure(show="" if cur else "•")

    def _pick_backup(self):
        path = filedialog.askopenfilename(
            title="Виберіть резервну копію Пром",
            filetypes=[("Excel файли", "*.xlsx *.xls"), ("Всі файли", "*.*")]
        )
        if path:
            self.backup_path.set(path)
            base = os.path.splitext(path)[0]
            if not self.out_path.get():
                self.out_path.set(base + "-new-products.xlsx")

    def _pick_out(self):
        path = filedialog.asksaveasfilename(
            title="Зберегти як",
            defaultextension=".xlsx",
            filetypes=[("Excel файли", "*.xlsx"), ("Всі файли", "*.*")]
        )
        if path:
            self.out_path.set(path)

    def _open_folder(self):
        path = self.out_path.get()
        if path and os.path.exists(path):
            folder = os.path.dirname(os.path.abspath(path))
            if sys.platform == "win32":
                os.startfile(folder)

    # ── Лог ──────────────────────────────────

    def _log(self, text, color=None):
        self.log_box.configure(state="normal")
        tag = None
        if color:
            tag = f"c_{color.replace('#','')}"
            self.log_box.tag_configure(tag, foreground=color)
        self.log_box.insert("end", text + "\n", tag or "")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _log_clear(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    # ── Прогрес-бар ──────────────────────────

    def _pb_start(self):
        self._pb_anim = 0
        self._pb_dir  = 1
        self._pb_tick()

    def _pb_tick(self):
        if not self._running:
            self.pb_fill.place(x=0, y=0, height=6, width=0)
            return
        w     = self.pb_bg.winfo_width()
        bar_w = max(w // 4, 60)
        self._pb_anim += self._pb_dir * 6
        if self._pb_anim >= w - bar_w:
            self._pb_dir = -1
        if self._pb_anim <= 0:
            self._pb_dir  = 1
            self._pb_anim = 0
        self.pb_fill.place(x=self._pb_anim, y=0, height=6, width=bar_w)
        self.after(20, self._pb_tick)

    # ── Налаштування ─────────────────────────

    def _settings_path(self):
        return os.path.join(os.path.expanduser("~"), ".profetch_xls_gen.ini")

    def _save_settings(self):
        try:
            with open(self._settings_path(), "w", encoding="utf-8") as f:
                f.write(f"token={self.prom_token.get()}\n")
                f.write(f"backup={self.backup_path.get()}\n")
                f.write(f"out={self.out_path.get()}\n")
        except Exception:
            pass

    def _load_settings(self):
        try:
            p = self._settings_path()
            if not os.path.exists(p):
                return
            with open(p, encoding="utf-8") as f:
                for line in f:
                    k, _, v = line.strip().partition("=")
                    if k == "token":
                        self.prom_token.set(v)
                    elif k == "backup":
                        self.backup_path.set(v)
                    elif k == "out":
                        self.out_path.set(v)
        except Exception:
            pass

    # ── Запуск ───────────────────────────────

    def _start(self):
        backup = self.backup_path.get().strip()
        token  = self.prom_token.get().strip()
        out    = self.out_path.get().strip()

        if not backup:
            messagebox.showwarning("Увага", "Виберіть файл резервної копії Пром")
            return
        if not os.path.exists(backup):
            messagebox.showwarning("Увага", "Файл резервної копії не знайдено")
            return
        if not token:
            messagebox.showwarning("Увага", "Введіть API-токен Пром.юа")
            return
        if not out:
            messagebox.showwarning("Увага", "Вкажіть шлях для збереження результату")
            return

        self._save_settings()
        self._log_clear()
        self._running = True
        self.run_btn.configure(state="disabled", text="⏳  Обробка...")
        self.open_btn.configure(state="disabled")
        for s in [self.stat_backup, self.stat_prom, self.stat_new]:
            s.configure(text="…", fg=TEXT_DIM)
        self._pb_start()

        threading.Thread(target=self._worker,
                         args=(backup, token, out), daemon=True).start()

    def _worker(self, backup, token, out):
        def log(msg, color=None):
            self.after(0, lambda: self._log(msg, color))

        try:
            # 1. Пром API
            log("⟳  Підключення до Пром.юа API...")
            prom_skus, prom_ext_ids = get_prom_products(
                token, lambda m: log(m)
            )
            log(f"✓  Пром: {len(prom_skus)} артикулів завантажено", SUCCESS)

            # 2. Резервна копія
            log("")
            log("⟳  Читання резервної копії...")
            headers, backup_rows, sku_idx, ext_idx = read_prom_backup(
                backup, lambda m: log(m)
            )
            total = len(backup_rows)
            log(f"✓  Резервна копія: {total} товарів, {len(headers)} колонок", SUCCESS)

            # 3. Порівняння
            log("")
            log("⟳  Порівняння з Пром...")
            new_rows, already = find_new_products(
                backup_rows, sku_idx, ext_idx,
                prom_skus, prom_ext_ids,
                lambda m: log(m)
            )
            log(f"✓  Порівняння завершено:", SUCCESS)
            log(f"   У резервній копії: {total}")
            log(f"   Вже є на Пром:     {already}")
            log(f"   НОВИХ товарів:     {len(new_rows)}",
                SUCCESS if new_rows else TEXT_DIM)

            # Оновлення статистики
            self.after(0, lambda: [
                self.stat_backup.configure(text=str(total),        fg=TEXT),
                self.stat_prom.configure(  text=str(already),      fg=TEXT),
                self.stat_new.configure(   text=str(len(new_rows)),
                                           fg=SUCCESS if new_rows else TEXT_DIM)
            ])

            # 4. Генерація XLS
            if new_rows:
                log("")
                log("⟳  Генерація XLS файлу...")
                os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
                generate_prom_xls(headers, new_rows, out, lambda m: log(m))

                log("")
                log("─" * 44, TEXT_DIM)
                log("ЩО РОБИТИ ДАЛІ:", ACCENT)
                log(f"  1. Відкрийте отриманий файл в Excel — перевірте дані", TEXT)
                log(f"  2. Пром → Товари → Імпорт позицій", TEXT)
                log(f"  3. Завантажте: {os.path.basename(out)}", TEXT)
                log(f"  4. БЕЗ галочки «Тільки оновлення»", WARNING)
                log(f"  5. Галочки: Ціна, Наявність, Характеристики", TEXT)
                log(f"  6. «Почати імпорт»", TEXT)
                log(f"  → {len(new_rows)} нових товарів додадуться з характеристиками", SUCCESS)
                log("─" * 44, TEXT_DIM)

                self.after(0, lambda: self.open_btn.configure(state="normal"))
            else:
                log("")
                log("✅  Нових товарів немає — всі вже є на Пром!", SUCCESS)

        except requests.exceptions.HTTPError as e:
            log(f"✗  HTTP помилка: {e}", DANGER)
            if "401" in str(e):
                log("   Перевірте API-токен Пром", DANGER)
        except requests.exceptions.ConnectionError:
            log("✗  Немає з'єднання з інтернетом", DANGER)
        except Exception as e:
            log(f"✗  Помилка: {e}", DANGER)
            import traceback
            log(traceback.format_exc(), DANGER)
        finally:
            self._running = False
            self.after(0, lambda: self.run_btn.configure(
                state="normal", text="▶   ЗНАЙТИ НОВІ ТОВАРИ → XLS"
            ))


# ──────────────────────────────────────────────
#  ТОЧКА ВХОДУ
# ──────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
