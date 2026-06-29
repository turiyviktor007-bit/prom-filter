#!/usr/bin/env python3
"""
Profetch — Нові товари → Пром XLS
Джерело: mrseller-1780.xml
Генерує XLS у форматі Пром з перекладеними характеристиками.
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os, sys, time, re
import xml.etree.ElementTree as ET
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── КОЛЬОРИ ──────────────────────────────────
BG, BG2, BG3 = "#0f1117", "#1a1d27", "#22263a"
ACCENT, SUCCESS, WARNING, DANGER = "#4f8ef7", "#3ecf8e", "#f7b731", "#fc5c65"
TEXT, TEXT_DIM, BORDER = "#e8eaf6", "#6b7280", "#2d3250"
FONT_TITLE = ("Segoe UI", 13, "bold")
FONT_LABEL = ("Segoe UI", 9)
FONT_SMALL = ("Segoe UI", 8)
FONT_MONO  = ("Consolas", 9)
FONT_BIG   = ("Segoe UI", 22, "bold")
FONT_BTN   = ("Segoe UI", 10, "bold")

# ── ЗАГОЛОВКИ (51 колонка, українська) ───────
PROM_HEADERS = [
    "Код_товару","Назва_позиції","Назва_позиції_укр",
    "Пошукові_запити","Пошукові_запити_укр","Опис","Опис_укр",
    "Тип_товару","Ціна","Валюта","Одиниця_виміру",
    "Мінімальний_обсяг_замовлення","Оптова_ціна","Мінімальне_замовлення_опт",
    "Посилання_зображення","Наявність","Кількість","Номер_групи",
    "Назва_групи","Посилання_підрозділу","Можливість_поставки",
    "Термін_поставки","Спосіб_пакування","Спосіб_пакування_укр",
    "Унікальний_ідентифікатор","Ідентифікатор_товару",
    "Ідентифікатор_підрозділу","Ідентифікатор_групи","Виробник",
    "Країна_виробник","Знижка","ID_групи_різновидів",
    "Особисті_нотатки","Продукт_на_сайті","Термін_дії_знижки_від",
    "Термін_дії_знижки_до","Ціна_від","Ярлик","HTML_заголовок",
    "HTML_заголовок_укр","HTML_опис","HTML_опис_укр",
    "Код_маркування_(GTIN)","Номер_пристрою_(MPN)","Вага,кг",
    "Ширина,см","Висота,см","Довжина,см","Де_знаходиться_товар",
    "Товар_в_ProSale","Чому_товар_не_в_ProSale",
]

# ── ХАРАКТЕРИСТИКИ: порядок і назви ──────────
# Рівно 21 позиція як в еталоні Прому
CHAR_ORDER = [
    "Колір",                              # 1
    "Матеріал",                           # 2
    "Форм-фактор",                        # 3
    "Особливості",                        # 4
    "Стан",                               # 5
    "Призначення",                        # 6
    "Особливість кольору",                # 7
    "Сумісність з бездротовою зарядкою",  # 8
    "Тип застежки",                       # 9
    "Візерунки і принти",                 # 10
    "Модель",                             # 11
    "Модель телефону",                    # 12
    "Сумісність з",                       # 13
    "Підтримка MagSafe",                  # 14
    "", "", "", "", "", "", "",            # 15-21 пусто
]

# ── СЛОВНИКИ ПЕРЕКЛАДУ (рос→укр) з XML Прому ─
# Кожен paramid MrSeller → (uk_char_name, {ru_val: uk_val})

# Колір (paramid=4846 в Пром, в MrSeller назва "Цвет")
COLORS_RU_UK = {
    "Красный":"Червоний","Розовый":"Рожевий","Оранжевый":"Помаранчевий",
    "Золотистый":"Золотистий","Бежевый":"Бежевий","Коричневый":"Коричневий",
    "Персиковый":"Персиковий","Желтый":"Жовтий","Песочный":"Пісочний",
    "Зелёный":"Зелений","Оливковый":"Оливковий","Хаки":"Хакі",
    "Бирюзовый":"Бірюзовий","Синий":"Синій","Голубой":"Блакитний",
    "Фиолетовый":"Фіолетовий","Сиреневый":"Бузковий","Черный":"Чорний",
    "Белый":"Білий","Серый":"Сірий","Сливовый":"Сливовий",
    "Бордовый":"Бордовий","Разные цвета":"Різні кольори","Прозрачный":"Прозорий",
    "Салатовый":"Салатовий","Серебряный":"Срібний","Серебристый":"Сріблястий",
    "Коралловый":"Кораловий","Темно-синий":"Темно-синій","Пудровый":"Пудровий",
    "Фуксия":"Фуксія","Мятный":"М'ятний","Фисташковый":"Фісташковий",
    "Лиловый":"Ліловий","Светло-голубой":"Світло-блакитний",
    "Розовое золото":"Рожеве золото","Изумрудный":"Смарагдовий",
    "Пурпурный":"Пурпурний","Светло-розовый":"Світло-рожевий",
    "Терракотовый":"Теракотовий","Графит":"Графіт","Золотой":"Золотий",
    "Темно-зелёный":"Темно-зелений","Ателье светлый":"Ательє світлий",
    "Midnight":"Чорний","Black":"Чорний","Bleached Bone":"Кремовий",
    "Clear":"Прозорий","Ash":"Сірий","Violet":"Фіолетовий",
    "Navy":"Темно-синій","Red":"Червоний","Blue":"Синій",
    "Green":"Зелений","Gray":"Сірий","White":"Білий",
    "Orange":"Помаранчевий","Purple":"Фіолетовий","Pink":"Рожевий",
    "Titanium":"Сірий","Olive":"Оливковий",
}

# Матеріал (paramid=644)
MATERIALS_RU_UK = {
    "Пластик":"Пластик","Пластмасса":"Пластмаса","Дерево":"Дерево",
    "Резина":"Гума","Металл":"Метал","Ткань":"Тканина",
    "Натуральная кожа":"Натуральна шкіра","Искусственная кожа":"Штучна шкіра",
    "Стекло":"Скло","Полиэстер":"Поліестер","Джинс":"Джинс",
    "Силикон":"Силікон","ТПУ (термопластичный полиуретан)":"ТПУ (термопластичний поліуретан)",
    "Полиуретан":"Поліуретан","Карбон":"Карбон","Микрофибра":"Мікрофібра",
    "Мех искусственный":"Хутро штучне","Мех натуральный":"Хутро натуральне",
    "Войлок":"Повсть","Бамбуковое волокно":"Бамбукове волокно",
    "Поликарбонат":"Полікарбонат","Алькантара":"Алькантара",
    "Арамид":"Арамід","Силикон + пластик":"Силікон + пластик",
    "Кожа + силикон":"Шкіра + силікон","Искусственная кожа+силикон":"Штучна шкіра+силікон",
    "Микротвил":"Мікротвіл","Гипс, пластик":"Гіпс, пластик",
    "Глина, пластик":"Глина, пластик","Дерево, металл":"Дерево, метал",
    # Додаткові з фіду
    "Силикон + Пластик":"Силікон + пластик","Полікарбонат":"Полікарбонат",
    "Кожа + микрофибра":"Шкіра + мікрофібра","Термополиуретан":"ТПУ (термопластичний поліуретан)",
    "Клеевой слой":"Клейовий шар",
}

# Форм-фактор (paramid=7042)
FORMFACTOR_RU_UK = {
    "Сменная задняя панель":"Змінна задня панель","Бампер":"Бампер",
    "Панель (Накладка на корпус)":"Панель (Накладка на корпус)","Панель":"Панель (Накладка на корпус)",
    "Чехол-книжка":"Чохол-книжка","Чехол-флип":"Чохол-фліп",
    "Чехол-футляр":"Чохол-футляр",
    "Плоское":"Плоске","Защитное стекло":"Захисне скло",
    "Чехол-кошелек":"Гаманець-чохол",
}

# Особливості (paramid=11867) — multiselect
FEATURES_RU_UK = {
    "Водонепроницаемый":"Водонепроникний","Крепление на пояс":"Кріплення на пояс",
    "Для занятий спортом":"Для занять спортом","Защитная плёнка в комплекте":"В комплекті захисна плівка",
    "Противоударный":"Протиударний","Защитное стекло":"Захисне скло",
    "Шейный шнурок":"Шийний шнурок","Ремешок на руку":"Ремінець на руку",
    "Рисунок":"Малюнок","Пылезащищенный":"Пилозахищений",
    "Влагозащищенный":"Вологозахищений","Чехол-аккумулятор":"Чохол-акумулятор",
    "Магнитный":"Магнітний","Чехол-клавиатура":"Чохол-клавіатура",
    "Функция подставки":"Функція підставки",
    "Карман для пластиковых карт и купюр":"Кишеня для пластикових карт і купюр",
    "Рифленая текстура":"Рифлена текстура","Покрытие soft touch":"Покриття soft touch",
    "Кольцо-держатель вращающееся":"Кільце-тримач обертове",
    "Защита камеры":"Захист камери","MagSafe":"MagSafe",
    "Усиленные борты чехла":"Посилені борти чохла","С магнитом":"З магнітом",
    # Додаткові
    "Поддержка бесконтактных платежей":"Підтримка безконтактних платежів",
    "Возможность использовать беспроводное ЗУ":"Можливість використовувати бездротовий ЗП",
    "Защита всего корпуса":"Захист всього корпусу",
    "Возможность использования магнитных держателей":"Можливість використання магнітних тримачів",
    "Сменные кнопки в комплекте":"Змінні кнопки в комплекті",
    "Покрытие anti-shock":"Покриття anti-shock",
    "Для занятий спортом":"Для занять спортом",
}

# Стан (paramid=10006)
STATE_RU_UK = {
    "Новое":"Новий","Б/У":"Вживані","Новий":"Новий",
}

# Призначення (paramid=5527)
PURPOSE_RU_UK = {
    "Для телефона":"Для телефону","Для плеера":"Для плеєра",
    "Для планшета":"Для планшета",
}

# Особливість кольору (paramid=11301)
COLOR_FEATURE_RU_UK = {
    "Матовый":"Матовий","Глянцевый":"Глянсовий","Прозрачный":"Прозорий",
}

# Тип застежки (paramid=11901)
CLASP_RU_UK = {
    "Липучка":"Липучка","Молния":"Блискавка","Кнопки":"Кнопки",
    "Без застежки":"Без застібки","Пуговицы":"Гудзики",
    "Затягивается ременем/шнурком":"Затягується ременем/шнурком",
    "Магнит":"Магніт","Защелка":"Клямка",
}

# Візерунки і принти (paramid=2719)
PRINTS_RU_UK = {
    "Горошек":"Горошок","Геометрический узор":"Геометричний візерунок",
    "Клетка":"Клітинка","Без узоров и принтов":"Без візерунків і принтів",
    "Абстрактный принт":"Абстрактний принт","Другое":"Інше",
    "Полоска":"Смужка","Текстура дерева":"Текстура дерева",
    "Надписи":"Написи","Однотонный":"Однотонний",
}

# Сумісність з (paramid=16752) — значення лишаємо як є (Apple, Samsung, Google)

# Підтримка MagSafe — bool (Да/Нет → Так/Ні)
BOOL_RU_UK = {"Да":"Так","Нет":"Ні","да":"Так","нет":"Ні",
               "Yes":"Так","No":"Ні","True":"Так","False":"Ні"}

# ── ДОВІДНИК ГРУП ────────────────────────────
PROM_GROUP_MAP = {
    "125025453":("Чехлы для смартфонов Apple",   "https://prom.ua/Chehly-dlya-telefonov","380230"),
    "125025490":("Чехлы для смартфонов Samsung", "https://prom.ua/Chehly-dlya-telefonov","380230"),
    "125025593":("Чехлы для смартфонов Google",  "https://prom.ua/Chehly-dlya-telefonov","380230"),
    "116950860":("Чехлы для мобильных телефонов","https://prom.ua/Chehly-dlya-telefonov","380230"),
    "128172640":("Стекла защитные","https://prom.ua/Zaschitnaya-plenka-dlya-mobilnyh-telefonov","5090318"),
    "131240127":("Наушники и гарнитуры","https://prom.ua/Naushniki-i-mikrofony","63715"),
    "131238983":("LG","https://prom.ua/Chehly-dlya-telefonov","380230"),
    "45546":("Корневая группа","","30602"),
}
DEFAULT_GROUP = "116950860"

BRAND_VALUEID_MAP = {
    "1053041":"125025453",  # Apple
    "1053225":"125025490",  # Samsung
    "1303680":"125025593",  # Google
}

# paramid → (uk_char_name, dict_to_use)
# None = пропускаємо, False = не перекладаємо (як є)
PARAMID_MAP = {
    "4846":  ("Колір",          COLORS_RU_UK),
    "23458": ("Колір",          COLORS_RU_UK),
    "644":   ("Матеріал",       MATERIALS_RU_UK),
    "23457": ("Матеріал",       MATERIALS_RU_UK),
    "195927":("Матеріал",       MATERIALS_RU_UK),
    "7042":  ("Форм-фактор",    FORMFACTOR_RU_UK),
    "27655": ("Форм-фактор",    FORMFACTOR_RU_UK),
    "195939":("Форм-фактор",    FORMFACTOR_RU_UK),
    "11867": ("Особливості",    FEATURES_RU_UK),
    "27723": ("Особливості",    FEATURES_RU_UK),
    "10006": ("Стан",           STATE_RU_UK),
    "5527":  ("Призначення",    PURPOSE_RU_UK),
    "23448": ("Призначення",    PURPOSE_RU_UK),
    "21806": ("Призначення",    PURPOSE_RU_UK),
    "11301": ("Особливість кольору", COLOR_FEATURE_RU_UK),
    "18923": ("Сумісність з бездротовою зарядкою", BOOL_RU_UK),
    "11901": ("Тип застежки",   CLASP_RU_UK),
    "2719":  ("Візерунки і принти", PRINTS_RU_UK),
    "19247": ("Підтримка MagSafe", BOOL_RU_UK),
    # Моделі телефону → збираємо окремо
    "105323": None,  # → model_parts
    "133649": None,  # → model_parts
    "243303": None,  # → model_parts
    # Сумісність з (бренд) → визначає групу
    "133625": None,
    "23451":  None,  # дублікат бренду
    # Країна виробник → колонка 30
    "98900":  None,
    # Габарити → колонки 45-48
    "48739":  None,  # вага
    "72949":  None,  # ширина
    "72950":  None,  # висота
    "72944":  None,  # глибина
    # EAN/MPN → колонки 43-44
    "26294":  None,
    "21784":  None,
    # Ігноруємо
    "23449":  None,
    "52257":  None,
    "2019":   None,
    "173580": None,
}


def translate(value, dictionary):
    """Перекладає значення за словником, якщо нема — лишає як є."""
    if not value:
        return ""
    # Пряме співпадіння
    if value in dictionary:
        return dictionary[value]
    # Для значень типу "Прозрачный: Pink" — беремо базовий колір
    if ":" in value:
        base = value.split(":")[0].strip()
        if base in dictionary:
            return dictionary[base]
    # Якщо не знайдено — лишаємо як є
    return value


def translate_multi(value, dictionary):
    """Перекладає мультизначення через |."""
    if not value:
        return ""
    parts = [v.strip() for v in value.split("|") if v.strip()]
    return "|".join(translate(p, dictionary) for p in parts)


# ── ЛОГІКА ───────────────────────────────────

def get_prom_skus(token, on_progress):
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    skus = set()
    limit, last_id, page = 100, None, 1
    while True:
        params = {"limit": limit, "status": "on_display,draft,deleted"}
        if last_id:
            params["last_id"] = last_id
        resp = requests.get("https://my.prom.ua/api/v1/products/list",
                            headers=headers, params=params, timeout=30)
        resp.raise_for_status()
        products = resp.json().get("products", [])
        if not products:
            break
        for p in products:
            sku = str(p.get("sku","")).strip()
            if sku:
                skus.add(sku)
        on_progress(f"   Сторінка {page}: {len(skus)} SKU")
        if len(products) < limit:
            break
        last_id = products[-1]["id"]
        page += 1
        time.sleep(0.4)
    return skus


def parse_feed(source, on_progress):
    on_progress("   Читання фіду...")
    if source.startswith("http"):
        resp = requests.get(source, timeout=60, headers={"User-Agent":"Mozilla/5.0"})
        resp.raise_for_status()
        content = resp.content
    else:
        with open(source,"rb") as f:
            content = f.read()

    root   = ET.fromstring(content)
    offers = root.findall(".//offer")
    on_progress(f"   Знайдено {len(offers)} офферів")

    products = []
    for o in offers:
        offer_id     = o.get("id","").strip()
        available    = o.get("available","false")
        vendor_code  = (o.findtext("vendorCode") or "").strip()
        name         = (o.findtext("name") or "").strip()
        name_ua      = (o.findtext("name_ua") or "").strip()
        desc         = (o.findtext("description") or "").strip()
        desc_ua      = (o.findtext("description_ua") or "").strip()
        price        = (o.findtext("price") or "").strip()
        currency     = (o.findtext("currencyId") or "UAH").strip()
        vendor       = (o.findtext("vendor") or "").strip()
        url          = (o.findtext("url") or "").strip()
        pictures     = [p.text.strip() for p in o.findall("picture") if p.text][:10]

        # Якщо нема укр — беремо рос
        if not name_ua:   name_ua  = name
        if not desc_ua:   desc_ua  = desc

        group_num    = DEFAULT_GROUP
        brand_compat = ""   # Apple / Samsung / Google для Сумісність з
        gtin         = ""
        mpn          = ""
        weight       = ""
        width        = ""
        height       = ""
        length       = ""
        country      = ""
        model_parts  = []
        chars        = {}   # uk_name → translated_value

        for p in o.findall("param"):
            paramid = p.get("paramid","").strip()
            pname   = (p.get("name") or "").strip()
            valueid = (p.get("valueid") or "").strip()
            pval    = (p.text or "").strip()
            if not pval:
                continue

            # Група з бренду
            if paramid == "133625":
                first_vid = valueid.split(",")[0].strip()
                if first_vid in BRAND_VALUEID_MAP and group_num == DEFAULT_GROUP:
                    group_num = BRAND_VALUEID_MAP[first_vid]
                # Значення для Сумісність з
                brand_name = pval.strip()
                if brand_name in ("Apple","Samsung","Google") and not brand_compat:
                    brand_compat = brand_name
                continue

            if paramid in ("23451","23456"):  # дублікат бренду/совмест
                continue

            if paramid == "98900":  # Країна
                country = pval; continue

            if paramid == "26294":  # EAN → GTIN
                digits = re.sub(r"[^0-9]","",pval)
                if digits.isdigit() and len(digits) in (8,12,13,14):
                    gtin = digits
                elif not mpn:
                    mpn = pval
                continue

            if paramid == "21784":  # MPN
                if not mpn: mpn = pval
                continue

            if paramid == "48739": weight = pval; continue
            if paramid == "72949": width  = pval; continue
            if paramid == "72950": height = pval; continue
            if paramid == "72944": length = pval; continue

            if paramid in ("105323","133649","243303"):  # Моделі
                parts = [v.strip() for v in pval.split(",") if v.strip()]
                model_parts.extend(parts)
                continue

            if paramid in PARAMID_MAP:
                mapping = PARAMID_MAP[paramid]
                if mapping is None:
                    continue
                uk_name, dictionary = mapping
                # Особливості — multiselect (розділювач кома або |)
                if uk_name == "Особливості":
                    sep = "|" if "|" in pval else ","
                    parts = [v.strip() for v in pval.split(sep) if v.strip()]
                    translated = "|".join(translate(v, dictionary) for v in parts)
                elif uk_name in ("Сумісність з бездротовою зарядкою","Підтримка MagSafe"):
                    translated = translate(pval, BOOL_RU_UK)
                else:
                    translated = translate(pval, dictionary)
                if uk_name and translated:
                    chars[uk_name] = translated
                continue

            # Решта — пропускаємо (не включаємо до характеристик)

        # Модель телефону через |
        model_tel = ""
        if model_parts:
            seen, unique = set(), []
            for m in model_parts:
                if m not in seen:
                    seen.add(m); unique.append(m)
            model_tel = "|".join(unique)

        # Сумісність з бездротовою зарядкою — якщо не знайдено, пусто
        if "Сумісність з бездротовою зарядкою" not in chars:
            if "MagSafe" in (chars.get("Особливості","") or ""):
                chars["Сумісність з бездротовою зарядкою"] = "Так"

        # Дефолти
        if "Стан" not in chars:
            chars["Стан"] = "Новий"
        if "Призначення" not in chars:
            chars["Призначення"] = "Для телефону"
        if "Візерунки і принти" not in chars:
            chars["Візерунки і принти"] = "Без візерунків і принтів"
        if "Форм-фактор" not in chars:
            chars["Форм-фактор"] = "Панель (Накладка на корпус)"

        products.append({
            "id": offer_id, "available": available,
            "vendor_code": vendor_code,
            "name": name, "name_ua": name_ua,
            "desc": desc, "desc_ua": desc_ua,
            "price": price, "currency": currency,
            "vendor": vendor, "url": url, "pictures": pictures,
            "group_num": group_num, "brand_compat": brand_compat,
            "gtin": gtin, "mpn": mpn, "country": country,
            "weight": weight, "width": width, "height": height, "length": length,
            "chars": chars, "model_tel": model_tel,
        })

    on_progress(f"   Розпарсено: {len(products)} товарів")
    return products


def find_new(products, prom_skus, only_available, on_progress):
    on_progress("   Порівняння з Пром...")
    new_items, already, skipped = [], 0, 0
    for p in products:
        if only_available and p["available"] != "true":
            skipped += 1; continue
        if p["id"] in prom_skus:
            already += 1
        else:
            new_items.append(p)
    return new_items, already, skipped


def generate_xls(new_items, output_path, on_progress):
    on_progress("   Формування XLS...")

    # 51 базових + 21 × 3 = 114 колонок рівно
    headers = list(PROM_HEADERS)
    for _ in range(21):
        headers += ["Назва_Характеристики",
                    "Одиниця_виміру_Характеристики",
                    "Значення_Характеристики"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Export Products Sheet"

    hfont  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hfill  = PatternFill("solid", start_color="1F4E79")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = hfont, hfill, halign
    ws.row_dimensions[1].height = 32
    dfont = Font(name="Arial", size=9)

    for idx, p in enumerate(new_items):
        avail    = "!" if p["available"] == "true" else "-"
        prosale  = "Так" if p["available"] == "true" else "Ні"
        prosale_why = "" if p["available"] == "true" else "Товару немає в наявності"
        grp_num  = p["group_num"]
        grp_name, grp_addr, grp_subid = PROM_GROUP_MAP.get(
            grp_num, ("","https://prom.ua/Chehly-dlya-telefonov","380230"))
        vc = p["vendor_code"] or p["id"]

        row = [
            p["id"],                    # 1  Код_товару
            p["name"],                  # 2  Назва_позиції
            p["name_ua"],               # 3  Назва_позиції_укр
            "", "",                     # 4-5 Пошукові запити
            p["desc"],                  # 6  Опис
            p["desc_ua"],               # 7  Опис_укр
            "r",                        # 8  Тип_товару
            p["price"],                 # 9  Ціна
            p["currency"],              # 10 Валюта
            "шт.",                      # 11 Одиниця_виміру
            "", "", "",                 # 12-14
            ", ".join(p["pictures"]),   # 15 Посилання_зображення
            avail,                      # 16 Наявність
            "",                         # 17 Кількість
            grp_num,                    # 18 Номер_групи
            grp_name,                   # 19 Назва_групи
            grp_addr,                   # 20 Посилання_підрозділу
            "", "", "", "",             # 21-24
            "",                         # 25 Унікальний_ідентифікатор (пусто)
            vc,                         # 26 Ідентифікатор_товару
            grp_subid,                  # 27 Ідентифікатор_підрозділу
            "",                         # 28 Ідентифікатор_групи
            p["vendor"],                # 29 Виробник
            p["country"],               # 30 Країна_виробник
            "", "", "",                 # 31-33
            p["url"],                   # 34 Продукт_на_сайті
            "", "",                     # 35-36
            "-",                        # 37 Ціна_від
            "",                         # 38 Ярлик
            "", "", "", "",             # 39-42 HTML
            p["gtin"],                  # 43 GTIN
            p["mpn"],                   # 44 MPN
            p["weight"],                # 45 Вага,кг
            p["width"],                 # 46 Ширина,см
            p["height"],                # 47 Висота,см
            p["length"],                # 48 Довжина,см
            "Київ",                     # 49 Де_знаходиться_товар
            prosale,                    # 50 Товар_в_ProSale
            prosale_why,                # 51 Чому_товар_не_в_ProSale
        ]

        # 21 характеристика в фіксованому порядку
        chars = dict(p["chars"])
        # Підставляємо Сумісність з і Модель телефону
        if p["brand_compat"]:
            chars["Сумісність з"] = p["brand_compat"]
        if p["model_tel"]:
            chars["Модель телефону"] = p["model_tel"]

        for char_name in CHAR_ORDER:
            if not char_name:
                row += ["", "", ""]
            else:
                val = chars.get(char_name, "")
                row += [char_name if val else "", "", val]

        # None → пусто і виправляємо пусті пари характеристик
        row = ["" if v is None else v for v in row]
        # Якщо назва характеристики пуста — очищаємо всю трійку
        for ci in range(51, len(row), 3):
            if ci+2 < len(row) and not str(row[ci]).strip():
                row[ci] = ""; row[ci+1] = ""; row[ci+2] = ""

        ws.append(row)
        for cell in ws[idx + 2]:
            cell.font = dfont

    # Ширини колонок
    for col, w in {1:12,2:50,3:50,6:60,7:60,9:10,
                   15:80,16:6,18:12,19:30,29:18,43:15,44:15}.items():
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    ws.freeze_panes = "A2"
    wb.save(output_path)
    kb = os.path.getsize(output_path)/1024
    on_progress(f"   Збережено: {os.path.basename(output_path)} ({kb:.0f} КБ)")


# ── GUI ──────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Profetch — Нові товари → Пром XLS")
        self.geometry("700x740")
        self.minsize(640,660)
        self.configure(bg=BG)
        self.feed_src   = tk.StringVar()
        self.prom_token = tk.StringVar()
        self.out_path   = tk.StringVar()
        self.only_avail = tk.BooleanVar(value=True)
        self._running   = False
        self._build_ui()
        self._load_settings()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=BG, pady=18)
        hdr.pack(fill="x", padx=24)
        tk.Label(hdr,text="⬡",font=("Segoe UI",20),fg=ACCENT,bg=BG).pack(side="left")
        tk.Label(hdr,text="  PROFETCH",font=FONT_TITLE,fg=TEXT,bg=BG).pack(side="left")
        tk.Label(hdr,text="  Нові товари → XLS",font=("Segoe UI",10),
                 fg=TEXT_DIM,bg=BG).pack(side="left",pady=3)
        self._sep()
        self._section("1  Фід MrSeller (файл або URL)")
        tk.Label(self,text="https://hub.tatet.net/mrseller-1780.xml  або виберіть XML файл",
                 font=FONT_SMALL,fg=TEXT_DIM,bg=BG).pack(anchor="w",padx=28,pady=(0,4))
        r1=tk.Frame(self,bg=BG); r1.pack(fill="x",padx=24,pady=(0,8))
        e1=self._entry(r1,self.feed_src,50); e1.pack(side="left",fill="x",expand=True,padx=(0,8))
        self._btn_small(r1,"Огляд…",self._pick_feed).pack(side="left")
        self._section("2  API-токен Пром.юа")
        tk.Label(self,text="Пром → Налаштування → Управління API-токенами",
                 font=FONT_SMALL,fg=TEXT_DIM,bg=BG).pack(anchor="w",padx=28,pady=(0,4))
        r2=tk.Frame(self,bg=BG); r2.pack(fill="x",padx=24,pady=(0,8))
        self.tok_entry=self._entry(r2,self.prom_token,50,show="•")
        self.tok_entry.pack(side="left",fill="x",expand=True,padx=(0,8))
        self._btn_small(r2,"👁",lambda:self.tok_entry.configure(
            show="" if self.tok_entry.cget("show") else "•")).pack(side="left")
        self._section("3  Зберегти результат як")
        r3=tk.Frame(self,bg=BG); r3.pack(fill="x",padx=24,pady=(0,8))
        e3=self._entry(r3,self.out_path,50); e3.pack(side="left",fill="x",expand=True,padx=(0,8))
        self._btn_small(r3,"Огляд…",self._pick_out).pack(side="left")
        cf=tk.Frame(self,bg=BG); cf.pack(fill="x",padx=24,pady=(2,0))
        tk.Checkbutton(cf,text="Тільки товари в наявності (available=true)",
                       variable=self.only_avail,font=FONT_SMALL,bg=BG,fg=TEXT_DIM,
                       selectcolor=BG3,activebackground=BG,activeforeground=TEXT,
                       bd=0,highlightthickness=0).pack(anchor="w")
        self._sep()
        self.run_btn=tk.Button(self,text="▶   ЗНАЙТИ НОВІ ТОВАРИ → XLS",
            font=FONT_BTN,bg=ACCENT,fg="#ffffff",
            activebackground="#7c5cfc",activeforeground="#ffffff",
            bd=0,cursor="hand2",padx=28,pady=12,command=self._start)
        self.run_btn.pack(fill="x",padx=24,pady=(12,4))
        pb=tk.Frame(self,bg=BG,height=6); pb.pack(fill="x",padx=24,pady=(0,8))
        pb.pack_propagate(False)
        self.pb_bg=tk.Frame(pb,bg=BG3,height=6); self.pb_bg.pack(fill="both",expand=True)
        self.pb_fill=tk.Frame(self.pb_bg,bg=ACCENT,height=6,width=0)
        self.pb_fill.place(x=0,y=0,height=6)
        self._pb_anim,self._pb_dir=0,1
        tk.Label(self,text="Журнал",font=FONT_SMALL,fg=TEXT_DIM,bg=BG).pack(anchor="w",padx=24)
        self.log_box=tk.Text(self,font=FONT_MONO,bg=BG2,fg=TEXT,
                             insertbackground=TEXT,bd=0,pady=10,padx=12,
                             relief="flat",state="disabled",height=8,selectbackground=BG3)
        self.log_box.pack(fill="both",expand=True,padx=24,pady=(4,0))
        sf=tk.Frame(self,bg=BG3,pady=14); sf.pack(fill="x",padx=24,pady=8)
        self.stat_feed=self._stat_cell(sf,"У фіді","—")
        self.stat_prom=self._stat_cell(sf,"Вже на Пром","—")
        self.stat_new =self._stat_cell(sf,"НОВИХ → XLS","—",color=SUCCESS)
        for c in range(3): sf.columnconfigure(c,weight=1)
        self.open_btn=tk.Button(self,text="📂  Відкрити папку з результатом",
            font=FONT_SMALL,bg=BG2,fg=TEXT_DIM,activebackground=BG3,activeforeground=TEXT,
            bd=0,cursor="hand2",padx=16,pady=8,command=self._open_folder,state="disabled")
        self.open_btn.pack(pady=(0,16))

    def _sep(self):
        tk.Frame(self,bg=BORDER,height=1).pack(fill="x",padx=24,pady=8)
    def _section(self,text):
        tk.Label(self,text=text,font=("Segoe UI",9,"bold"),fg=ACCENT,bg=BG
                 ).pack(anchor="w",padx=24,pady=(12,4))
    def _entry(self,parent,var,width=40,show=None):
        kw=dict(textvariable=var,font=FONT_LABEL,bg=BG3,fg=TEXT,
                insertbackground=TEXT,bd=0,relief="flat",
                highlightthickness=1,highlightbackground=BORDER,
                highlightcolor=ACCENT,width=width)
        if show: kw["show"]=show
        e=tk.Entry(parent,**kw); self._enable_paste(e); return e
    def _enable_paste(self,entry):
        def do_paste(event=None):
            try:
                text=self.clipboard_get()
                try: entry.delete("sel.first","sel.last")
                except tk.TclError: pass
                entry.insert("insert",text)
            except tk.TclError: pass
            return "break"
        def do_copy(event=None):
            try: self.clipboard_clear(); self.clipboard_append(entry.selection_get())
            except tk.TclError: pass
            return "break"
        def do_cut(event=None):
            do_copy()
            try: entry.delete("sel.first","sel.last")
            except tk.TclError: pass
            return "break"
        def select_all(event=None):
            entry.select_range(0,"end"); entry.icursor("end"); return "break"
        menu=tk.Menu(entry,tearoff=0,bg=BG3,fg=TEXT,
                     activebackground=ACCENT,activeforeground="#ffffff",bd=0)
        menu.add_command(label="Вставити",command=do_paste)
        menu.add_command(label="Копіювати",command=do_copy)
        menu.add_command(label="Вирізати",command=do_cut)
        menu.add_separator()
        menu.add_command(label="Виділити все",command=select_all)
        def show_menu(event):
            try: menu.tk_popup(event.x_root,event.y_root)
            finally: menu.grab_release()
        entry.bind("<Button-3>",show_menu)
        for seq,fn in [("<Control-v>",do_paste),("<Control-V>",do_paste),
                       ("<Control-c>",do_copy),("<Control-C>",do_copy),
                       ("<Control-x>",do_cut),("<Control-X>",do_cut),
                       ("<Control-a>",select_all),("<Control-A>",select_all)]:
            entry.bind(seq,fn)
        def kh(event):
            if event.state&0x4:
                if event.keycode==86: return do_paste()
                if event.keycode==67: return do_copy()
                if event.keycode==88: return do_cut()
                if event.keycode==65: return select_all()
        entry.bind("<Key>",kh)
    def _btn_small(self,parent,text,cmd):
        return tk.Button(parent,text=text,font=FONT_SMALL,bg=BG3,fg=TEXT_DIM,
                         activebackground=BORDER,activeforeground=TEXT,
                         bd=0,cursor="hand2",padx=10,pady=6,relief="flat",
                         command=cmd,highlightthickness=1,highlightbackground=BORDER)
    def _stat_cell(self,parent,label,value,color=TEXT):
        frame=tk.Frame(parent,bg=BG3)
        frame.grid(row=0,column=len(parent.winfo_children())-1,sticky="nsew",padx=6)
        tk.Label(frame,text=label,font=FONT_SMALL,fg=TEXT_DIM,bg=BG3).pack(pady=(8,0))
        val=tk.Label(frame,text=value,font=FONT_BIG,fg=color,bg=BG3); val.pack(pady=(0,8))
        return val
    def _pick_feed(self):
        path=filedialog.askopenfilename(title="Виберіть XML фід",
            filetypes=[("XML","*.xml"),("Всі файли","*.*")])
        if path:
            self.feed_src.set(path)
            if not self.out_path.get():
                self.out_path.set(os.path.splitext(path)[0]+"-new.xlsx")
    def _pick_out(self):
        path=filedialog.asksaveasfilename(title="Зберегти як",defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx"),("Всі файли","*.*")])
        if path: self.out_path.set(path)
    def _open_folder(self):
        path=self.out_path.get()
        if path and os.path.exists(path) and sys.platform=="win32":
            os.startfile(os.path.dirname(os.path.abspath(path)))
    def _log(self,text,color=None):
        self.log_box.configure(state="normal")
        if color:
            tag=f"c{color.replace('#','')}"
            self.log_box.tag_configure(tag,foreground=color)
            self.log_box.insert("end",text+"\n",tag)
        else:
            self.log_box.insert("end",text+"\n")
        self.log_box.see("end"); self.log_box.configure(state="disabled")
    def _log_clear(self):
        self.log_box.configure(state="normal"); self.log_box.delete("1.0","end")
        self.log_box.configure(state="disabled")
    def _pb_start(self):
        self._pb_anim,self._pb_dir=0,1; self._pb_tick()
    def _pb_tick(self):
        if not self._running:
            self.pb_fill.place(x=0,y=0,height=6,width=0); return
        w=self.pb_bg.winfo_width(); bar_w=max(w//4,60)
        self._pb_anim+=self._pb_dir*6
        if self._pb_anim>=w-bar_w: self._pb_dir=-1
        if self._pb_anim<=0: self._pb_dir,self._pb_anim=1,0
        self.pb_fill.place(x=self._pb_anim,y=0,height=6,width=bar_w)
        self.after(20,self._pb_tick)
    def _settings_path(self):
        return os.path.join(os.path.expanduser("~"),".profetch_new.ini")
    def _save_settings(self):
        try:
            with open(self._settings_path(),"w",encoding="utf-8") as f:
                f.write(f"token={self.prom_token.get()}\n")
                f.write(f"feed={self.feed_src.get()}\n")
                f.write(f"out={self.out_path.get()}\n")
        except Exception: pass
    def _load_settings(self):
        try:
            p=self._settings_path()
            if not os.path.exists(p): return
            with open(p,encoding="utf-8") as f:
                for line in f:
                    k,_,v=line.strip().partition("=")
                    if k=="token": self.prom_token.set(v)
                    elif k=="feed": self.feed_src.set(v)
                    elif k=="out": self.out_path.set(v)
        except Exception: pass
    def _start(self):
        feed=self.feed_src.get().strip(); token=self.prom_token.get().strip()
        out=self.out_path.get().strip()
        if not feed: messagebox.showwarning("Увага","Вкажіть фід або URL"); return
        if not feed.startswith("http") and not os.path.exists(feed):
            messagebox.showwarning("Увага","Файл фіду не знайдено"); return
        if not token: messagebox.showwarning("Увага","Введіть API-токен Пром.юа"); return
        if not out: messagebox.showwarning("Увага","Вкажіть куди зберегти результат"); return
        self._save_settings(); self._log_clear(); self._running=True
        self.run_btn.configure(state="disabled",text="⏳  Обробка...")
        self.open_btn.configure(state="disabled")
        for s in (self.stat_feed,self.stat_prom,self.stat_new):
            s.configure(text="…",fg=TEXT_DIM)
        self._pb_start()
        threading.Thread(target=self._worker,args=(feed,token,out),daemon=True).start()
    def _worker(self,feed,token,out):
        def log(msg,color=None): self.after(0,lambda:self._log(msg,color))
        try:
            log("⟳  Підключення до Пром.юа API...")
            prom_skus=get_prom_skus(token,lambda m:log(m))
            log(f"✓  Пром: {len(prom_skus)} SKU",SUCCESS)
            log(""); log("⟳  Читання фіду MrSeller...")
            products=parse_feed(feed,lambda m:log(m))
            log(f"✓  Фід: {len(products)} товарів",SUCCESS)
            log(""); log("⟳  Порівняння...")
            new_items,already,skipped=find_new(
                products,prom_skus,self.only_avail.get(),lambda m:log(m))
            total=len(products)
            log(f"✓  Результат:",SUCCESS)
            log(f"   У фіді:        {total}")
            if skipped: log(f"   Не в наявності (пропущено): {skipped}",WARNING)
            log(f"   Вже на Пром:   {already}")
            log(f"   НОВИХ:         {len(new_items)}",
                SUCCESS if new_items else TEXT_DIM)
            self.after(0,lambda:[
                self.stat_feed.configure(text=str(total),fg=TEXT),
                self.stat_prom.configure(text=str(already),fg=TEXT),
                self.stat_new.configure(text=str(len(new_items)),
                    fg=SUCCESS if new_items else TEXT_DIM)])
            if new_items:
                log(""); log("⟳  Генерація XLS...")
                os.makedirs(os.path.dirname(os.path.abspath(out)),exist_ok=True)
                generate_xls(new_items,out,lambda m:log(m))
                log(""); log("─"*44,TEXT_DIM)
                log("ЩО РОБИТИ ДАЛІ:",ACCENT)
                log("  1. Відкрийте файл в Excel — перевірте",TEXT)
                log("  2. Пром → Товари → Імпорт позицій",TEXT)
                log(f"  3. Завантажте: {os.path.basename(out)}",TEXT)
                log("  4. БЕЗ галочки «Тільки оновлення»",WARNING)
                log("  5. «Почати імпорт»",TEXT)
                log(f"  → {len(new_items)} нових товарів",SUCCESS)
                log("─"*44,TEXT_DIM)
                self.after(0,lambda:self.open_btn.configure(state="normal"))
            else:
                log(""); log("✅  Нових товарів немає — всі вже є на Пром!",SUCCESS)
        except requests.exceptions.HTTPError as e:
            log(f"✗  HTTP помилка: {e}",DANGER)
            if "401" in str(e): log("   Перевірте API-токен Пром",DANGER)
        except requests.exceptions.ConnectionError:
            log("✗  Немає з'єднання з інтернетом",DANGER)
        except ET.ParseError as e:
            log(f"✗  Помилка XML: {e}",DANGER)
        except Exception as e:
            log(f"✗  Помилка: {e}",DANGER)
            import traceback; log(traceback.format_exc(),DANGER)
        finally:
            self._running=False
            self.after(0,lambda:self.run_btn.configure(
                state="normal",text="▶   ЗНАЙТИ НОВІ ТОВАРИ → XLS"))

if __name__ == "__main__":
    App().mainloop()
